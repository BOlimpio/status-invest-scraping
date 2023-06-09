import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def scrape_fii_info(stock_code):
    url = f"https://statusinvest.com.br/fundos-imobiliarios/{stock_code}"
    headers = {
      'accept': '*/*',
      'accept-language': 'en-US,en;q=0.9,pt;q=0.8',
      'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
      'x-requested-with': 'XMLHttpRequest',
      'sec-ch-ua': 'Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111',
      'sec-ch-ua-mobile': '?0',
      'sec-ch-ua-platform': '"Linux"',
      'sec-fetch-dest': 'document',
      'sec-fetch-mode': 'navigate',
      'sec-fetch-site': 'same-origin',
      'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'
    }
    
    response = requests.get(url, headers=headers)
    messages = []

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Extracting required information from the website
        try:
            price_element = soup.find(text='Valor atual')
            if price_element:
                price = price_element.find_next(class_='value').text.strip()
            else:
                messages.append(f"Could not find 'Valor atual' field for stock {stock_code}")
                price = None
            
            p_vp_element = soup.find(text='P/VP')
            if p_vp_element:
                p_vp = p_vp_element.find_next(class_='value').text.strip()
            else:
                messages.append(f"Could not find 'P/VP' field for stock {stock_code}")
                p_vp = None
            
            dy_element = soup.find(title='Dividend Yield com base nos últimos 12 meses')
            if dy_element:
                dy = dy_element.find_next(class_='value').text.strip('%')
            else:
                messages.append(f"Could not find 'Dividend Yield' field for stock {stock_code}")
                dy = None
            
            dy_cagr_3_element = soup.select_one('h3.title:-soup-contains("DY CAGR") + strong.value')
            if dy_cagr_3_element:
                dy_cagr_3_value = dy_cagr_3_element.text.strip()
            else:
                messages.append(f"Could not find 'DY CAGR 3 anos' field for stock {stock_code}")
                dy_cagr_3_value = None

            # Extracting new fields for 'Último rendimento'
            ultimo_rendimento_element = soup.find(text='Último rendimento')
            if ultimo_rendimento_element:
                ultimo_rendimento_value = ultimo_rendimento_element.find_next(class_='value').text.strip()
                ultimo_rendimento_percentage = ultimo_rendimento_element.find_next(class_='sub-value').text.strip('%')

                ultimo_rendimento_subinfo = ultimo_rendimento_element.find_next(class_='sub-info')
                ultima_cotacao_base = ultimo_rendimento_subinfo.find_next(text='Cotação base').find_next(class_='sub-value').text.strip()
                ultima_data_pagamento = ultimo_rendimento_subinfo.find_next(text='Data Pagamento').find_next(class_='sub-value').text.strip()
            else:
                messages.append(f"Could not find 'Último rendimento' field for stock {stock_code}")
                ultimo_rendimento_value = None
                ultimo_rendimento_percentage = None
                ultima_cotacao_base = None
                ultima_data_pagamento = None

            # Extracting new fields for 'Próximo Rendimento'
            proximo_rendimento_element = soup.find(text='Próximo Rendimento')
            if proximo_rendimento_element:
                proximo_rendimento_value = proximo_rendimento_element.find_next(class_='value').text.strip()
                proximo_rendimento_percentage = proximo_rendimento_element.find_next(class_='sub-value').text.strip('%')

                proximo_rendimento_subinfo = proximo_rendimento_element.find_next(class_='sub-info')
                proxima_cotacao_base = proximo_rendimento_subinfo.find_next(text='Cotação base').find_next(class_='sub-value').text.strip()
                proxima_data_pagamento = proximo_rendimento_subinfo.find_next(text='Data Pagamento').find_next(class_='sub-value').text.strip()
            else:
                messages.append(f"Could not find 'Próximo Rendimento' field for stock {stock_code}")
                proximo_rendimento_value = None
                proximo_rendimento_percentage = None
                proxima_cotacao_base = None
                proxima_data_pagamento = None
            
            return {'Code': stock_code, 'Price': price, 'P/VP': p_vp, 'DY %': dy, 'DY CAGR 3 anos': dy_cagr_3_value,
                    'Último Rendimento': ultimo_rendimento_value, 'Último Rendimento %': ultimo_rendimento_percentage, 'Ultima Cotação base' : ultima_cotacao_base, 'Ultima data de pagamento' : ultima_data_pagamento,
                    'Próximo Rendimento': proximo_rendimento_value, 'Próximo Rendimento %': proximo_rendimento_percentage, 'Próxima Cotação base' : proxima_cotacao_base, 'Próxima data de pagamento' : proxima_data_pagamento,
                    'Messages': messages}
        
        except AttributeError as e:
            messages.append(f"Failed to extract information for stock {stock_code}. Error: {e}")
            return {'Code': stock_code, 'Messages': messages}
    else:
        messages.append(f"Failed to retrieve data for stock {stock_code}. Status code: {response.status_code}, Response: {response.text}")
        return {'Code': stock_code, 'Messages': messages}

def save_to_excel(data, filename):
    error_messages = []
    
    try:
        df = pd.DataFrame(data)
        df.to_excel(filename, index=False)

        # Open the workbook
        workbook = openpyxl.load_workbook(filename)

        # Get the active sheet
        sheet = workbook.active

        # Find the index of the "P/VP" column
        p_vp_index = df.columns.get_loc("P/VP")

        # Apply styling to "P/VP" column based on the rule
        for row in sheet.iter_rows(min_row=2, min_col=p_vp_index+1, max_col=p_vp_index+1):
            for cell in row:
                if cell.value is not None:
                    p_vp_value = float(cell.value.replace(',', '.'))
                    if p_vp_value > 1:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    elif p_vp_value < 1:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        workbook.save(filename)

    except Exception as e:
        error_messages.append(f"Failed to save data to Excel file. Error: {e}")
    
    return error_messages
